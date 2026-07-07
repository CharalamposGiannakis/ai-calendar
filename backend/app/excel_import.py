from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class ExcelExtractionError(ValueError):
    pass


@dataclass(frozen=True)
class ExtractedExcelRow:
    row_index: int
    values: list[Any]
    source_locator_json: str
    raw_text: str
    raw_data_json: str


@dataclass(frozen=True)
class ExtractedWorksheetRows:
    worksheet_name: str
    rows: list[ExtractedExcelRow]


def _compact_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=True, separators=(",", ":"))


def _cell_value_for_json(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def _row_has_value(values: list[Any]) -> bool:
    return any(value is not None and value != "" for value in values)


def extract_first_visible_worksheet_rows(workbook_path: Path) -> ExtractedWorksheetRows:
    try:
        workbook = load_workbook(
            filename=workbook_path,
            read_only=True,
            data_only=True,
        )
    except Exception as error:
        raise ExcelExtractionError("Workbook extraction failed.") from error

    try:
        worksheet = next(
            (sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"),
            None,
        )
        if worksheet is None:
            raise ExcelExtractionError("Workbook has no visible worksheets.")

        extracted_rows: list[ExtractedExcelRow] = []
        for row_index, row_values in enumerate(
            worksheet.iter_rows(values_only=True),
            start=1,
        ):
            values = [_cell_value_for_json(value) for value in row_values]
            if not _row_has_value(values):
                continue

            raw_text = _compact_json(values)
            source_locator = {
                "worksheet": worksheet.title,
                "row": row_index,
            }
            extracted_rows.append(
                ExtractedExcelRow(
                    row_index=row_index,
                    values=values,
                    source_locator_json=_compact_json(source_locator),
                    raw_text=raw_text,
                    raw_data_json=raw_text,
                )
            )

        return ExtractedWorksheetRows(
            worksheet_name=worksheet.title,
            rows=extracted_rows,
        )
    except ExcelExtractionError:
        raise
    except Exception as error:
        raise ExcelExtractionError("Workbook extraction failed.") from error
    finally:
        workbook.close()
